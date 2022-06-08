import openpyxl

# url 저장
fpath = "/Users/minxi/Github_repo/Web_scraper/test_crawling/append_data.xlsx"

# 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 시트 선택하기
ws = wb["위코드"]

# 데이터 수정
ws["A3"] = "조성호"
ws["B3"] = "프론트엔드"

# 데이터 저장
wb.save(fpath)