import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet("위코드")

ws["A1"] = "이름"
ws["B1"] = "분야"

ws["A2"] = "이민석"
ws["B2"] = "백엔드"

wb.save("/Users/minxi/Github_repo/Web_scraper/test_crawling/append_data.xlsx")